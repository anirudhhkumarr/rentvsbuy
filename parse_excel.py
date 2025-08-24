#!/usr/bin/env python3
"""
Read actual values from Excel sheet to use as test cases
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def read_excel_values():
    """Read actual values from the Excel sheet for testing"""
    
    # Load the workbook
    wb = load_workbook('Household Finances.xlsx', data_only=True)  # data_only=True to get calculated values
    
    # Get the RentBuy sheet
    ws = wb['RentBuy']
    
    print("Reading actual values from Excel sheet...")
    print("=" * 60)
    
    # Read key parameter values
    print("KEY PARAMETERS:")
    print("-" * 30)
    
    # Read parameters from specific cells
    params = {
        'B7': ws['B7'].value,  # Mortgage rate
        'B8': ws['B8'].value,  # Mortgage points
        'B13': ws['B13'].value,  # Tax + maintenance
        'B17': ws['B17'].value,  # Rent increase
        'B18': ws['B18'].value,  # Property reassessment rate
        'B20': ws['B20'].value,  # Home return
        'B21': ws['B21'].value,  # Stock return
        'B22': ws['B22'].value,  # Inflation
    }
    
    for cell, value in params.items():
        print(f"{cell}: {value}")
    
    print("\n" + "=" * 60)
    print("COLUMN LABELS FROM ROW 2:")
    print("=" * 60)
    
    # Read column labels from row 2
    column_labels = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value is not None:
            # Convert column number to Excel column reference
            if col <= 26:
                col_letter = chr(64 + col)
            elif col <= 52:
                col_letter = 'A' + chr(64 + col - 26)
            elif col <= 78:
                col_letter = 'B' + chr(64 + col - 52)
            else:
                col_letter = 'C' + chr(64 + col - 78)
            
            column_labels[col_letter] = cell.value
            print(f"{col_letter}: {cell.value}")
    
    print("\n" + "=" * 60)
    print("REFERENCED CELLS AND THEIR VALUES:")
    print("=" * 60)
    
    # Read labels from column A and values from column B for key rows
    key_cells = {
        'B3': {'label_cell': 'A3', 'value_cell': 'B3'},
        'B5': {'label_cell': 'A5', 'value_cell': 'B5'},
        'B8': {'label_cell': 'A8', 'value_cell': 'B8'},
        'B9': {'label_cell': 'A9', 'value_cell': 'B9'},
        'B10': {'label_cell': 'A10', 'value_cell': 'B10'},
        'B16': {'label_cell': 'A16', 'value_cell': 'B16'},
        'C12': {'label_cell': 'A12', 'value_cell': 'C12'}
    }
    
    for cell_ref, cells in key_cells.items():
        label = ws[cells['label_cell']].value
        value = ws[cells['value_cell']].value
        if value is not None:
            if isinstance(value, (int, float)):
                if abs(value) >= 1000000:
                    formatted_value = f"${value/1000000:.1f}M"
                elif abs(value) >= 1000:
                    formatted_value = f"${value/1000:.0f}K"
                else:
                    formatted_value = f"${value:.0f}"
            else:
                formatted_value = str(value)
        else:
            formatted_value = "N/A"
        
        print(f"{cell_ref} ({label}): {formatted_value}")

    print("\n" + "=" * 60)
    print("C12 MORTGAGE PAYMENT FORMULA:")
    print("=" * 60)
    
    # Load formulas workbook
    wb_formulas = load_workbook('Household Finances.xlsx', data_only=False)
    ws_formulas = wb_formulas['RentBuy']
    
    # Check C12 formula specifically
    c12_formula = ws_formulas['C12'].value
    c12_value = ws['C12'].value
    print(f"C12 Value: {c12_value}")
    print(f"C12 Formula: {c12_formula}")
    
    # Also check if there are any other payment-related formulas
    for row in range(10, 15):
        for col in ['B', 'C', 'D']:
            cell_ref = f"{col}{row}"
            formula = ws_formulas[cell_ref].value
            value = ws[cell_ref].value
            label = ws[f"A{row}"].value
            if formula and isinstance(formula, str) and formula.startswith('='):
                print(f"{cell_ref} ({label}): {value} | Formula: {formula}")

    print("\n" + "=" * 60)
    print("ROW 3 - INITIAL VALUES AND FORMULAS:")
    print("=" * 60)
    
    # Use the actual column labels from row 2
    columns = column_labels
    
    # Read initial values and formulas from row 3
    print("Row 3 Values:")
    row3_data = []
    for col in columns.keys():
        cell_value = ws[f'{col}{3}'].value
        if cell_value is None:
            row3_data.append("N/A")
        elif isinstance(cell_value, (int, float)):
            if abs(cell_value) >= 1000000:
                row3_data.append(f"${cell_value/1000000:.1f}M")
            elif abs(cell_value) >= 1000:
                row3_data.append(f"${cell_value/1000:.0f}K")
            else:
                row3_data.append(f"${cell_value:.0f}")
        else:
            row3_data.append(str(cell_value))
    
    print("Row 3 | " + " | ".join(row3_data))
    
    print("\nRow 3 Formulas:")
    row3_formulas = {}
    wb_formulas = load_workbook('Household Finances.xlsx', data_only=False)  # data_only=False to get formulas
    ws_formulas = wb_formulas['RentBuy']
    for col in columns.keys():
        formula = ws_formulas[f'{col}{3}'].value
        if formula and isinstance(formula, str) and formula.startswith('='):
            row3_formulas[col] = formula
            print(f"{col}3: {formula}")
        else:
            print(f"{col}3: {formula} (no formula)")
    
    print("\n" + "=" * 60)
    print("ROW 4 - FIRST CALCULATED VALUES AND FORMULAS:")
    print("=" * 60)
    
    # Read first calculated values and formulas from row 4
    print("Row 4 Values:")
    row4_data = []
    for col in columns.keys():
        cell_value = ws[f'{col}{4}'].value
        if cell_value is None:
            row4_data.append("N/A")
        elif isinstance(cell_value, (int, float)):
            if abs(cell_value) >= 1000000:
                row4_data.append(f"${cell_value/1000000:.1f}M")
            elif abs(cell_value) >= 1000:
                row4_data.append(f"${cell_value/1000:.0f}K")
            else:
                row4_data.append(f"${cell_value:.0f}")
        else:
            row4_data.append(str(cell_value))
    
    print("Row 4 | " + " | ".join(row4_data))
    
    print("\nRow 4 Formulas:")
    row4_formulas = {}
    for col in columns.keys():
        formula = ws_formulas[f'{col}{4}'].value
        if formula and isinstance(formula, str) and formula.startswith('='):
            row4_formulas[col] = formula
            print(f"{col}4: {formula}")
        else:
            print(f"{col}4: {formula} (no formula)")
    
    print("\n" + "=" * 60)
    print("COMPLETE ANALYSIS - ROWS 3 TO 10:")
    print("=" * 60)
    
    # Read values and formulas for rows 3-10 to understand the complete pattern
    for row in range(3, 11):
        print(f"\n{'='*20} ROW {row} {'='*20}")
        
        # Print values in organized format
        print("VALUES:")
        for col in columns.keys():
            cell_value = ws[f'{col}{row}'].value
            if cell_value is None:
                formatted_value = "N/A"
            elif isinstance(cell_value, (int, float)):
                if abs(cell_value) >= 1000000:
                    formatted_value = f"${cell_value/1000000:.3f}M"
                elif abs(cell_value) >= 1000:
                    formatted_value = f"${cell_value/1000:.1f}K"
                else:
                    formatted_value = f"${cell_value:.2f}"
            else:
                formatted_value = str(cell_value)
            print(f"  {col}{row} ({columns[col]}): {formatted_value}")
        
        # Print formulas
        print("FORMULAS:")
        for col in columns.keys():
            formula = ws_formulas[f'{col}{row}'].value
            if formula and isinstance(formula, str) and formula.startswith('='):
                print(f"  {col}{row}: {formula}")
            else:
                print(f"  {col}{row}: {formula} (direct value)")
        
        print("-" * 50)
    
    print("\n" + "=" * 60)
    print("FORMULA COMPARISON:")
    print("=" * 60)
    
    # Read formulas for key cells to verify our understanding
    print("Key formulas from Excel:")
    print("-" * 30)
    
    # Check formulas for year 11 (row 11)
    key_formulas = {
        'E11': ws_formulas['E11'].value,
        'F11': ws_formulas['F11'].value,
        'G11': ws_formulas['G11'].value,
        'H11': ws_formulas['H11'].value,
        'I11': ws_formulas['I11'].value,
        'J11': ws_formulas['J11'].value,
        'K11': ws_formulas['K11'].value,
        'L11': ws_formulas['L11'].value,
        'M11': ws_formulas['M11'].value,
        'N11': ws_formulas['N11'].value,
        'P11': ws_formulas['P11'].value,
        'Q11': ws_formulas['Q11'].value,
        'R11': ws_formulas['R11'].value,
        'S11': ws_formulas['S11'].value,
        'T11': ws_formulas['T11'].value,
        'U11': ws_formulas['U11'].value,
        'V11': ws_formulas['V11'].value,
        'W11': ws_formulas['W11'].value,
        'X11': ws_formulas['X11'].value,
        'Z11': ws_formulas['Z11'].value
    }
    
    for cell, formula in key_formulas.items():
        if formula and formula.startswith('='):
            print(f"{cell}: {formula}")
    
    print("\n" + "=" * 60)
    print("SUMMARY:")
    print("=" * 60)
    
    # Calculate some key metrics
    year_10_buy_real = ws['N10'].value
    year_10_rent_real = ws['X10'].value
    year_20_buy_real = ws['N20'].value
    year_20_rent_real = ws['X20'].value
    
    if year_10_buy_real and year_10_rent_real:
        print(f"Year 0 (Row 10): Buy Real: {year_10_buy_real:,.0f}, Rent Real: {year_10_rent_real:,.0f}")
        print(f"Year 0 Premium: {year_10_rent_real - year_10_buy_real:,.0f}")
    
    if year_20_buy_real and year_20_rent_real:
        print(f"Year 10 (Row 20): Buy Real: {year_20_buy_real:,.0f}, Rent Real: {year_20_rent_real:,.0f}")
        print(f"Year 10 Premium: {year_20_rent_real - year_20_buy_real:,.0f}")

if __name__ == "__main__":
    read_excel_values()
